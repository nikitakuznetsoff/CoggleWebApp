import networkx as nx
from openpyxl.styles import PatternFill
import  openpyxl


# Группа методов для преобразования данных о вершинах в граф networkx
# Начало алгоритма для списка деревьев
def transform_into_graph(node):
    passed_nodes = []
    graph = nx.DiGraph()
    for obj in node:
        passed_nodes.append(obj['_id'])
        graph.add_node(obj['_id'], text=obj['text'], offset=obj['offset'])
        #graph.add_node(obj['_id'])
        transform_into_graph_rec(obj, passed_nodes, graph)
    return graph


# Начало алгоритма для конкретного дерева
def transform_into_graph_algo(node):
    passed_nodes = []
    graph = nx.DiGraph()
    graph.add_node(node['_id'], text=node['text'], offset=node['offset'])
    transform_into_graph_rec(node, passed_nodes, graph)
    return graph


# Метод предназначенный для рекурсивного использования
def transform_into_graph_rec(node, passed_nodes, graph):
    if node['children']:
        for obj in node['children']:
            try:
                if passed_nodes.index(obj['_id']):
                    continue
            except Exception:
                passed_nodes.append(obj['_id'])
                graph.add_node(obj['_id'], text=obj['text'], offset=obj['offset'])
                graph.add_edge(node['_id'], obj['_id'], color=obj['colour'])
                transform_into_graph_rec(obj, passed_nodes, graph)
    return graph


# Алгоритм для экспортирования в таблицу матрицы
def print_matrix(sheet, mass):
    wb = openpyxl.Workbook()
    ws = wb.active

    if len(mass) == 0:
        sheet.cell(1, 1, "Проблема с входными данными")
    else:
        for i in range(2, len(mass[0]) + 2):
            sheet.cell(1, i, i - 1)

        for i in range(2, len(mass) + 2):
            sheet.cell(i, 1, i - 1)

        for i in range(0, len(mass)):
            for j in range(0, len(mass[0])):
                sheet.cell(i + 2, j + 2, mass[i][j])



# Алгоритм для экспортирования в таблицу метрик
def print_metrics(output_sheet, metrics_diagram_1, metrics_diagram_2):
    names_metrics = ["max_height", "count_nodes", "count_first_layer_branches", "images", "avg_node_text_len"]
    output_sheet.cell(1, 3, "Первая карта")
    output_sheet.cell(1, 4, "Вторая карта")
    for i in range(2, 7):
        output_sheet.cell(i, 2, names_metrics[i-2])
        output_sheet.cell(i, 3, metrics_diagram_1[names_metrics[i-2]])
        output_sheet.cell(i, 4, metrics_diagram_1[names_metrics[i-2]])


# Создание графа из списка максимальных элементов
def create_graph_form_list(arr):
    last_graph = nx.DiGraph()
    for obj in arr:
        last_graph.add_nodes_from(obj.nodes())
        last_graph.add_edges_from(obj.edges())
    return last_graph


# Чтение ИДшников референсных карт
def read_mm_ids(sheet, point="B2"):
    arr = []
    column = ord(point[0]) - 65
    str_row = point[1:]

    for row in sheet.iter_rows(None, int(str_row), None, column):
        try:
            qwe = row[column].value
            if row[column].value:
                arr.append(link_to_id(row[column].value.strip()))
            else:
                arr.append(None)
        except Exception:
            return arr
    return arr


# Возврат ИДшника из ссылки
def link_to_id(link):
    first = link.find("/diagram/")
    last = link.find("/t/")
    return link[first + 9:last:1]