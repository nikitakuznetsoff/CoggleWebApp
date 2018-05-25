import json
import requests
from requests.auth import HTTPBasicAuth
import webbrowser
import flask
from openpyxl import load_workbook
from flask import Flask, request, redirect, render_template
from werkzeug.utils import secure_filename
from functions import analysis_function as af
from functions import substructure_algo as sub_algo
from functions import graph_function as gf
from functions import subtree_isomorphism as si

app = Flask(__name__)
app.secret_key = 'some_secret'


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if coggle_user.access_token == "":
        coggle_user.authorization_token()
    if request.method == 'POST':
        form = request.form['form_algo']

        # Проверки загруки нужнных файлов
        if (form == "with_true") and (('file_input' not in request.files) or ('file_input_ref' not in request.files)):
            if 'file_input' not in request.files:
                flask.flash('Ошибка! Вы не загрузили файл', 'input')
            if 'file_input_ref' not in request.files:
                flask.flash('Ошибка! Вы не загрузили файл', 'input_ref')
            return redirect(request.url)

        if (form == "without_true") and ('file_input_ref' not in request.files):
            flask.flash('Ошибка! Вы не загрузили файл', 'input_ref')
            return redirect(request.url)

        # Делаем запросы для получения информации о файлах
        if form == "with_true":
            file = request.files['file_input']
        file_ref = request.files['file_input_ref']

        # Еще одни проверки на наличие файла
        if form == "with_true":
            if file.filename == '':
                flask.flash('Ошибка! Вы не выбрали файл!', 'input')
                return redirect(request.url)
        if file_ref.filename == '':
            flask.flash('Ошибка! Вы не выбрали файл!', 'input_ref')
            return redirect(request.url)


        # Считываем названия файлов
        filename_file = ""
        if form == "with_true":
            filename_file = secure_filename(file.filename)
            file.save(filename_file)
        filename_file_ref = secure_filename(file_ref.filename)
        file_ref.save(filename_file_ref)

        # Считываем данных о названий таблиц
        inf_for_read1 = request.form['inf_for_read1']
        inf_for_read2 = request.form['inf_for_read2']

        # Проверки названий таблиц
        if form == "with_true":
            if inf_for_read1 != "":
                wb = load_workbook(filename=filename_file)
                if not check_correct_tablename(wb, inf_for_read1):
                    flask.flash('Ошибка! В файле нет таблицы с таким названием!', 'inf_for_read1')
                    return redirect(request.url)
            if inf_for_read2 != "":
                wb = load_workbook(filename=filename_file_ref)
                if not check_correct_tablename(wb, inf_for_read2):
                    flask.flash('Ошибка! В файле нет таблицы с таким названием!', 'inf_for_read2')
                    return redirect(request.url)
        else:
            if inf_for_read2 != "":
                wb = load_workbook(filename=filename_file_ref)
                if not check_correct_tablename(wb, inf_for_read2):
                    flask.flash('Ошибка! В файле нет таблицы с таким названием!', 'inf_for_read2')
                    return redirect(request.url)

        # Считываем значения ячеек
        cell_for_read1 = request.form['cell_for_read1']
        cell_for_read2 = request.form['cell_for_read2']

        # Проверки на форматы ячеек
        if form == "with_true":
            if (cell_for_read1 != '') or (cell_for_read2 != ''):
                if (len(cell_for_read1) > 4 or len(cell_for_read2) > 4) or (not check_correct_cellname(cell_for_read1)) or (not check_correct_cellname(cell_for_read2)):
                    if (not check_correct_cellname(cell_for_read1)) or len(cell_for_read1) > 4:
                        flask.flash('Ошибка! Вы ввели ячеку некорректного формата!', 'cell_for_read1')
                    if (not check_correct_cellname(cell_for_read2)) or len(cell_for_read2) > 4:
                        flask.flash('Ошибка! Вы ввели ячеку некорректного формата!', 'cell_for_read2')
                    return redirect(request.url)
        else:
            if (not check_correct_cellname(cell_for_read2)) or len(cell_for_read2) > 4:
                flask.flash('Ошибка! Вы ввели ячеку некорректного формата!', 'cell_for_read2')
                return redirect(request.url)

        # Вызов метода, в котором вычисляются меры сходства и выводится информация
        if form == "with_true":
            status = True
            take_mm(filename_file_ref, inf_for_read2, cell_for_read2, status, filename_file, inf_for_read1, cell_for_read1)
        else:
            status = False
            take_mm(filename_file_ref, inf_for_read2, cell_for_read2, status)
    return render_template('index.html')


def take_mm(filename2, name2, cell2, status, filename1="", name1="", cell1=""):
    wb = load_workbook(filename2)
    if name2 != '':
        sheet = wb[name2]
    else:
        sheet = wb[wb.sheetnames[0]]

    if cell2 == '':
        arr_ids = gf.read_mm_ids(sheet)
    else:
        arr_ids = gf.read_mm_ids(sheet, cell2)
    arr_diagrams = create_arr_diagrams(arr_ids)

    if status:
        wb_first = load_workbook(filename1)
        if name1 != '':
            sheet_1 = wb_first[name1]
        else:
            sheet_1 = wb_first[wb_first.sheetnames[0]]

        if cell1 == '':
            arr_ids_first = gf.read_mm_ids(sheet_1)
        else:
            arr_ids_first = gf.read_mm_ids(sheet_1, cell1)
        arr_diagrams_first = create_arr_diagrams(arr_ids_first)

        mass = [[0] * len(arr_diagrams) for i in range(len(arr_diagrams_first))]
        for i in range(len(arr_diagrams_first)):
            for j in range(len(arr_diagrams)):
                if (arr_diagrams_first[i] == None) or (arr_diagrams[j] == None):
                    mass[i][j] = None
                else:
                    mass[i][j] = si.max_common_substree_rooted(arr_diagrams_first[i], arr_diagrams[j])
        ###
        output_sheet = wb.create_sheet('results')
        gf.print_matrix(output_sheet, mass)
        wb.save(filename2)
    else:
        mass = [[0] * len(arr_diagrams) for i in range(len(arr_diagrams))]
        for i in range(len(arr_diagrams)):
            for j in range(len(arr_diagrams)):
                if (arr_diagrams[i] == None) or (arr_diagrams[j] == None):
                    mass[i][j] = None
                else:
                    mass[i][j] = si.max_common_substree_rooted(arr_diagrams[i], arr_diagrams[j])
        ###
        output_sheet = wb.create_sheet('results')
        gf.print_matrix(output_sheet, mass)
        wb.save(filename2)


# Проверка на наличие конкретной таблицы в файле
def check_correct_tablename(wb, name):
    arr = wb.sheetnames
    try:
        if arr.index(name) != -1:
            return True
    except Exception:
        return False


# Проверка на кооректный формат ячейки для начала считывания
def check_correct_cellname(name):
    if name == "": return True
    if name[0].isupper() and name[1:].isdigit():
        return True
    return False


# Создание диаграммы из ID и проверка закрыта / открыта карта
def create_diagram(id):
    if id is None:
        return None
    diag = coggle_user.nodes(id)
    for obj in diag:
        if obj == 'error':
            return None
        break
    return diag


# Создание массива диаграмм из массива айдишников
def create_arr_diagrams(arr):
    new_arr = []
    for obj in arr:
        new_arr.append(create_diagram(obj))
    return new_arr


def information_for_algo(id_diagram):
    arr = {'diagram': '', 'graph': ''}
    arr['diagram'] = coggle_user.nodes(id_diagram)
    arr['graph'] = gf.transform_into_graph(arr['diagram'])
    return arr


# Класс для основных операций с API
class Coggle:
    def __init__(self, app_name, client_id, client_secret, redirect_uri):
        self.app_name = app_name
        self.client_id = client_id
        self.client_secret = client_secret
        self.redirect_uri = redirect_uri
        self.url_base = "https://coggle.it/"
        self.access_token = ""

    # Авторизация
    def authorization(self):
        params = {'response_type': "code", 'scope': "read", 'client_id': self.client_id, 'redirect_uri': self.redirect_uri}
        resp_auth = requests.get(self.url_base + 'dialog/authorize/', params=params)
        webbrowser.open_new_tab(resp_auth.url)
        app.run()

    def authorization_token(self):
        code = request.args.get('code')
        params = {"code": code, "grant_type": "authorization_code", "redirect_uri": self.redirect_uri}
        resp1 = requests.post(self.url_base + "token", auth=HTTPBasicAuth(client_id, self.client_secret), json=params)
        information_auth = json.loads(resp1.text)
        self.access_token = information_auth["access_token"]

    # Получение информации о всей диаграмме
    def diagram(self, id_diagram):
        params = {"access_token": self.access_token}
        url_diagram = "api/1/diagrams/"
        information_diagram = requests.get(self.url_base + url_diagram + id_diagram, params=params)
        diagram = json.loads(information_diagram.text)
        return diagram

    # Получение информации о вершинах
    def nodes(self, id_diagram):
        params = {"access_token": self.access_token}
        url_diagram = "api/1/diagrams/"
        information_nodes = requests.get(self.url_base + url_diagram + id_diagram + "/nodes", params=params)
        nodes = json.loads(information_nodes.text)
        return nodes


# Информация пользователя
app_name = "CourseWork"
client_id = "5a8ab7821c5b5b00010853f0"
client_secret = "52cc11494efb7b38cbcfd113fc8f6f67b0d09c5aaccce7a07f2a6c177a34c772"
redirect_uri = "http://localhost:5000"

coggle_user = Coggle(app_name, client_id, client_secret, redirect_uri)
coggle_user.authorization()

